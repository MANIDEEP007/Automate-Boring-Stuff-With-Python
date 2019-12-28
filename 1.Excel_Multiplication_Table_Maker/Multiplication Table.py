import openpyxl as op
from openpyxl.styles import Font
n = int(input("Enter an interger: "))
workbook_obj = op.Workbook() #Create Workbook
sheet_obj = workbook_obj[workbook_obj.sheetnames[0]] #Create Sheet Object
bfont = Font(bold=True)
for i in range(1,n+1):
	sheet_obj.cell(row=1,column=i+1).font = bfont
	sheet_obj.cell(row=1,column=i+1).value = i
	sheet_obj.cell(row=i+1,column=1).font = bfont
	sheet_obj.cell(row=i+1,column=1).value = i
for rows in range(1,n+1): #Calculating Multiplication_Table
	for cols in range(1,n+1):
		prod = sheet_obj.cell(row=rows+1,column=1).value*sheet_obj.cell(row=1,column=cols+1).value
		sheet_obj.cell(row=rows+1,column=cols+1).value = prod
workbook_obj.save("Multiplication_Table.xlsx") #Save Workbook
